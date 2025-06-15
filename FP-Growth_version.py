import pandas as pd
import zipfile
from io import BytesIO
from pathlib import Path
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from mlxtend.frequent_patterns import fpgrowth, association_rules
from scipy.sparse import csr_matrix
from sklearn.preprocessing import MultiLabelBinarizer
import logging
import time

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)

# === File Paths
historical_zip_path = Path("Historical_BOM.zip")
tobeadded_zip_path = Path("To_be_Added.zip")
critical_items_path = Path("KONE_Critical_Item.xlsx")
output_excel = Path("Updated_Historical_ARM.xlsx")
min_support = 0.035
max_files_for_fpgrowth = 2  # Maximum number of files to use for FP-Growth

# === Format Component
def format_component(code: str) -> str:
    code = str(code)
    match = re.search(r"(KM\d+)", code)
    return match.group(1) if match else code.split("/")[0]

# === Load BOMs from ZIP
def load_boms_from_zip(zip_path: Path, source_type: str) -> pd.DataFrame:
    logger.info(f"Loading BOMs from {zip_path}")
    start_time = time.time()
    records = []
    with zipfile.ZipFile(zip_path, 'r') as archive:
        for file_info in archive.infolist():
            if file_info.filename.endswith(".xlsx") and "__MACOSX" not in file_info.filename:
                logger.info(f"Processing file: {file_info.filename}")
                with archive.open(file_info) as file:
                    try:
                        df = pd.read_excel(BytesIO(file.read()), engine="openpyxl", dtype=str)
                        if "Component" in df.columns and "kmfg material" in df.columns:
                            cols = ["Component", "kmfg material"]
                            if "Description / TITLE" in df.columns:
                                cols.append("Description / TITLE")
                            df = df[cols].dropna()
                            df.columns = ["Component", "Material"] + (["Description / TITLE"] if len(cols) == 3 else [])
                            df["Component"] = df["Component"].apply(format_component)
                            df["Source_File"] = Path(file_info.filename).name
                            df["Source_Type"] = source_type
                            if "Description / TITLE" not in df.columns:
                                df["Description / TITLE"] = ""
                            records.append(df)
                    except Exception as e:
                        logger.error(f"Error processing {file_info.filename}: {str(e)}")
                        continue
    result = pd.concat(records, ignore_index=True) if records else pd.DataFrame(columns=["Component", "Material", "Description / TITLE", "Source_File", "Source_Type"])
    logger.info(f"Loaded {len(result)} records in {time.time() - start_time:.2f} seconds")
    return result

# === ARM Metric Calculation
def calculate_arm_metrics(df: pd.DataFrame) -> pd.DataFrame:
    logger.info("Calculating ARM metrics")
    start_time = time.time()
    unique = df[["Component", "Material", "Source_File"]].drop_duplicates()
    total_files = df["Source_File"].nunique()
    pair_counts = unique.groupby(["Component", "Material"]).size().reset_index(name="Count")
    comp_totals = unique.groupby("Component")["Source_File"].nunique().reset_index(name="Component_Total")
    result = pair_counts.merge(comp_totals, on="Component", how="left")
    result["Support"] = result["Count"] / total_files
    result["Confidence"] = result["Count"] / result["Component_Total"]
    result["Support_Confidence_Sum"] = result["Support"] + result["Confidence"]
    logger.info(f"ARM metrics calculated in {time.time() - start_time:.2f} seconds")
    return result.round(5)

# === Load Data
logger.info("Starting data loading process")
critical_df = pd.read_excel(critical_items_path, engine="openpyxl", dtype=str)
critical_items = set(critical_df["ItemID_KONE"].dropna().astype(str).str.strip().str.upper())
logger.info(f"Loaded {len(critical_items)} critical items")

hist_df = load_boms_from_zip(historical_zip_path, "Historical")
new_df = load_boms_from_zip(tobeadded_zip_path, "To_be_Added")

logger.info("Processing and combining data")
for df in [hist_df, new_df]:
    df["Component"] = df["Component"].str.strip().str.upper()
    df["Material"] = df["Material"].str.strip().str.upper()

combined_df = pd.concat([hist_df, new_df], ignore_index=True)
logger.info(f"Combined dataset size: {len(combined_df)} rows")
metrics_df = calculate_arm_metrics(combined_df)

# === Enrichment
desc_map = combined_df.drop_duplicates(subset=["Component", "Material"]).set_index(["Component", "Material"])["Description / TITLE"].to_dict()
source_map = combined_df.groupby(["Component", "Material"])["Source_File"].apply(lambda x: ", ".join(sorted(set(x)))).reset_index()

merged = metrics_df.merge(source_map, on=["Component", "Material"], how="left")
merged["Description / TITLE"] = merged.set_index(["Component", "Material"]).index.map(desc_map)
merged["Critical_Flag"] = merged["Component"].apply(lambda c: "Critical" if c in critical_items else "Safe")

# === Status Assignment
def assign_status(row):
    match = hist_df[
        (hist_df["Component"] == row["Component"]) &
        (hist_df["Material"] == row["Material"])
    ]
    if match.empty:
        return "New"
    elif row["Support"] < min_support:
        return "Rare"
    else:
        return "Not Rare"

merged["Status"] = merged.apply(assign_status, axis=1)

# === Apriori If-Then Rule Mining
logger.info("Starting Apriori rule mining")
start_time = time.time()

combined_df_copy = combined_df.copy()
combined_df_copy["Item"] = combined_df_copy["Component"] + "__" + combined_df_copy["Material"] + "__" + merged.set_index(["Component", "Material"]).loc[combined_df_copy.set_index(["Component", "Material"]).index]["Status"].values

# More aggressive filtering
logger.info("Applying aggressive filtering")
# 1. Filter by component frequency - keep only top 10%
component_counts = combined_df_copy["Component"].value_counts()
top_components = component_counts[component_counts >= component_counts.quantile(0.9)].index
logger.info(f"Selected top {len(top_components)} components")

# 2. Filter by material frequency - keep only top 10%
material_counts = combined_df_copy["Material"].value_counts()
top_materials = material_counts[material_counts >= material_counts.quantile(0.9)].index
logger.info(f"Selected top {len(top_materials)} materials")

# 3. Filter by file frequency - keep only top 10%
file_counts = combined_df_copy["Source_File"].value_counts()
top_files = file_counts[file_counts >= file_counts.quantile(0.9)].index
logger.info(f"Selected top {len(top_files)} files")

# 4. Apply all filters
combined_df_copy = combined_df_copy[
    (combined_df_copy["Component"].isin(top_components)) & 
    (combined_df_copy["Material"].isin(top_materials)) &
    (combined_df_copy["Source_File"].isin(top_files))
]

# Select top N files based on frequency
file_counts = combined_df_copy["Source_File"].value_counts()
selected_files = file_counts.head(max_files_for_fpgrowth).index
combined_df_copy = combined_df_copy[combined_df_copy["Source_File"].isin(selected_files)]
logger.info(f"Selected top {len(selected_files)} files for FP-Growth analysis")

# 5. Filter out rare items with higher threshold
item_counts = combined_df_copy["Item"].value_counts()
frequent_items = item_counts[item_counts >= min_support * 3 * len(combined_df_copy["Source_File"].unique())].index
combined_df_copy = combined_df_copy[combined_df_copy["Item"].isin(frequent_items)]
logger.info(f"After filtering: {len(combined_df_copy)} rows, {len(frequent_items)} unique items, {len(combined_df_copy['Source_File'].unique())} files")

# Create transaction matrix
transaction_df = pd.DataFrame({"File": combined_df_copy["Source_File"], "Item": combined_df_copy["Item"]})
transaction_basket = transaction_df.groupby("File")["Item"].apply(list)

# Convert to sparse matrix with reduced dimensions
logger.info("Converting to sparse matrix")
mlb = MultiLabelBinarizer()
basket_sparse = csr_matrix(mlb.fit_transform(transaction_basket))
basket_dummies = pd.DataFrame.sparse.from_spmatrix(basket_sparse, columns=mlb.classes_)
# Convert to boolean type for better performance
basket_dummies = basket_dummies.astype(bool)
logger.info(f"Final matrix shape: {basket_dummies.shape}")

# Run FP-Growth instead of Apriori
logger.info("Running FP-Growth algorithm")
frequent_itemsets = fpgrowth(basket_dummies, min_support=min_support * 3, use_colnames=True)
logger.info(f"Found {len(frequent_itemsets)} frequent itemsets")

rules = association_rules(frequent_itemsets, metric="confidence", min_threshold=0.8)
logger.info(f"Generated {len(rules)} rules in {time.time() - start_time:.2f} seconds")

rules = rules[(rules['antecedents'].apply(lambda x: len(x) == 1)) & (rules['consequents'].apply(lambda x: len(x) == 1))].copy()
rules["Antecedent"] = rules["antecedents"].apply(lambda x: list(x)[0])
rules["Consequent"] = rules["consequents"].apply(lambda x: list(x)[0])
rules[["Material", "Status"]] = rules["Consequent"].str.extract(r"__([^_]+)__([^_]+)$")
rules["Component"] = rules["Antecedent"].str.extract(r"^([^_]+)__")
final_rules = rules[["Component", "Material", "Status", "support", "confidence", "lift"]].rename(columns={
    "support": "Support", "confidence": "Confidence", "lift": "Lift"
})

# === Prepare Sheets
sheet1 = merged[["Component", "Material", "Count", "Support", "Confidence", "Source_File", "Description / TITLE", "Critical_Flag", "Status"]]
sheet2 = sheet1[sheet1["Status"] == "New"]
sheet3 = merged[["Component", "Material", "Count", "Support", "Confidence", "Support_Confidence_Sum"]].sort_values("Support_Confidence_Sum", ascending=False)
sheet4 = pd.concat([hist_df, new_df], ignore_index=True)

# === Write to Excel
logger.info("Writing results to Excel")
with pd.ExcelWriter(output_excel, engine="openpyxl") as writer:
    sheet1.to_excel(writer, sheet_name="All_BOMs_Combined", index=False)
    sheet2.to_excel(writer, sheet_name="To_be_Added_Only", index=False)
    sheet3.to_excel(writer, sheet_name="Material_Summary", index=False)
    sheet4.to_excel(writer, sheet_name="Merged_Sheet", index=False)
    final_rules.to_excel(writer, sheet_name="Apriori_IfThen", index=False)

# === Highlight new entries
wb = load_workbook(output_excel)
ws1 = wb["All_BOMs_Combined"]
status_col1 = [cell.value for cell in ws1[1]].index("Status") + 1
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
    if row[status_col1 - 1].value == "New":
        for cell in row:
            cell.fill = green_fill
ws4 = wb["Merged_Sheet"]
src_type_idx = [cell.value for cell in ws4[1]].index("Source_Type") + 1
for row in ws4.iter_rows(min_row=2, max_row=ws4.max_row):
    if row[src_type_idx - 1].value == "To_be_Added":
        for cell in row:
            cell.fill = green_fill
wb.save(output_excel)

# === Merge new files into historical zip and clear to_be_added
with zipfile.ZipFile(historical_zip_path, 'a') as hist_zip, zipfile.ZipFile(tobeadded_zip_path, 'r') as new_zip:
    for file_info in new_zip.infolist():
        if file_info.filename.endswith(".xlsx") and "__MACOSX" not in file_info.filename:
            hist_zip.writestr(file_info.filename, new_zip.read(file_info.filename))
with zipfile.ZipFile(tobeadded_zip_path, 'w') as clear_zip:
    pass

logger.info("✅ Final Excel generated with extended Apriori If-Then logic and full BOM analysis.")