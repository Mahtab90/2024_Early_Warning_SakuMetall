#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Jun 12 14:25:50 2025

@author: mahtab
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Final Accurate BOM Comparison with Clean Output and Correct Total Counts
Author: Mahtab Shahin
"""

import pandas as pd
import zipfile
from io import BytesIO
from pathlib import Path
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === Configuration ===
base_path = Path("/Users/mahtab/Desktop/AIRE/Input/")
output_path = Path("/Users/mahtab/Desktop/AIRE/Output/BOM_Compare_Professional_Final_AllSheetsWithCounts.xlsx")
support_threshold = 0.035

# === Files ===
historical_zip = base_path / "Historical_BOM.zip"
to_be_added_zip = base_path / "To_be_Added.zip"
updated_hist_zip = base_path / "Historical_BOM.zip"  # This will overwrite the original
critical_items_path = base_path / "KONE_Critical_Item.xlsx"

# === Helper functions ===
def format_component(code):
    code = str(code)
    match = re.search(r"(KM\d+)", code)
    return match.group(1) if match else code.split("/")[0]

def load_zip(zip_path, label):
    records = []
    with zipfile.ZipFile(zip_path, 'r') as z:
        for file in z.infolist():
            if file.filename.endswith(".xlsx") and "__MACOSX" not in file.filename:
                with z.open(file) as f:
                    df = pd.read_excel(BytesIO(f.read()), engine="openpyxl", dtype=str)
                    if "Component" in df.columns and "kmfg material" in df.columns:
                        cols = ["Component", "kmfg material"]
                        if "Description / TITLE" in df.columns:
                            cols.append("Description / TITLE")
                        df = df[cols].dropna()
                        df.columns = ["Component", "Material"] + (["Description / TITLE"] if len(cols) == 3 else [])
                        df["Component"] = df["Component"].apply(format_component).str.strip().str.upper()
                        df["Material"] = df["Material"].str.strip().str.upper()
                        df["Source_File"] = Path(file.filename).name
                        df["Source_Type"] = label
                        if "Description / TITLE" not in df.columns:
                            df["Description / TITLE"] = ""
                        records.append(df)
    return pd.concat(records, ignore_index=True)

def calculate_metrics(df):
    total_files = df["Source_File"].nunique()
    unique_files = df[["Component", "Material", "Source_File"]].drop_duplicates()
    comp_file_counts = unique_files.groupby(["Component", "Material"]).size().reset_index(name="File_Occurrence")
    component_totals = unique_files.groupby("Component")["Source_File"].nunique().reset_index(name="Component_Total")
    metrics = comp_file_counts.merge(component_totals, on="Component", how="left")
    metrics["Support"] = metrics["File_Occurrence"] / total_files
    metrics["Confidence"] = metrics["File_Occurrence"] / metrics["Component_Total"]
    metrics["Support_Confidence_Sum"] = metrics["Support"] + metrics["Confidence"]
    return metrics.round(5)

# === Load inputs ===
critical_items = set(pd.read_excel(critical_items_path)["ItemID_KONE"].dropna().str.strip().str.upper())
hist_df = load_zip(historical_zip, "Historical")
add_df = load_zip(to_be_added_zip, "To_be_Added")

# === Permanent merge step ===
existing_pairs = set(zip(hist_df["Component"], hist_df["Material"]))
new_df_filtered = add_df[~add_df.apply(lambda row: (row["Component"], row["Material"]) in existing_pairs, axis=1)]
updated_hist_df = pd.concat([hist_df, new_df_filtered], ignore_index=True)

# === Overwrite Historical_BOM.zip with merged content ===
with zipfile.ZipFile(updated_hist_zip, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
    for fname, group in updated_hist_df.groupby("Source_File"):
        buffer = BytesIO()
        group[["Component", "Material", "Description / TITLE"]].to_excel(buffer, index=False, engine="openpyxl")
        zf.writestr(fname, buffer.getvalue())

# === Assign Critical Flags ===
for df in [hist_df, add_df]:
    df["Critical_Flag"] = df["Component"].apply(lambda x: "Critical" if x in critical_items else "Safe")

# === Reference for tracking
ref_df = hist_df.copy()
ref_pairs = set(zip(ref_df["Component"], ref_df["Material"]))

# === Combined for analysis
combined_df = pd.concat([ref_df, add_df], ignore_index=True)
metrics_df = calculate_metrics(combined_df)

# === Sheet 1: Historical
hist_grouped = hist_df.drop_duplicates(subset=["Component", "Material"]).merge(metrics_df, on=["Component", "Material"], how="left")
hist_grouped["Status"] = hist_grouped["Support"].apply(lambda x: "Rare" if x < support_threshold else "Not Rare")

# === Sheet 2: New data + status
add_df["Is_New"] = add_df.apply(lambda row: (row["Component"], row["Material"]) not in ref_pairs, axis=1)
add_grouped = add_df.drop_duplicates(subset=["Component", "Material"]).merge(metrics_df, on=["Component", "Material"], how="left")
add_grouped["Status"] = add_grouped.apply(
    lambda row: "New" if row["Is_New"] else ("Rare" if row["Support"] < support_threshold else "Not Rare"),
    axis=1
)

# === Sheet 3: Merged
merged_df = pd.concat([hist_grouped, add_grouped], ignore_index=True)

# === Sheet 5: Total count with metrics
sheet5_df = combined_df.groupby(["Component", "Material"]).agg({
    "Component": "size",
    "Description / TITLE": "first",
    "Critical_Flag": "first"
}).rename(columns={"Component": "Total_Count"}).reset_index()
sheet5_df = sheet5_df.merge(metrics_df, on=["Component", "Material"], how="left")
sheet5_df["Status"] = sheet5_df["Support"].apply(lambda x: "Rare" if x < support_threshold else "Not Rare")

# === Export to Excel
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    hist_grouped.to_excel(writer, sheet_name="1_Historical", index=False)
    add_grouped.to_excel(writer, sheet_name="2_To_Be_Added", index=False)
    merged_df.to_excel(writer, sheet_name="3_Merged", index=False)
    metrics_df.to_excel(writer, sheet_name="4_Metrics", index=False)
    sheet5_df.to_excel(writer, sheet_name="5_Total_Count", index=False)

# === Highlight To_be_Added rows in green
wb = load_workbook(output_path)
ws = wb["3_Merged"]
src_idx = [cell.value for cell in ws[1]].index("Source_Type") + 1
green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    if "To_be_Added" in str(row[src_idx - 1].value):
        for cell in row:
            cell.fill = green
wb.save(output_path)

print(f"✅ Excel exported with counts → {output_path}")
print(f"✅ Historical_BOM.zip permanently updated at → {updated_hist_zip}")
